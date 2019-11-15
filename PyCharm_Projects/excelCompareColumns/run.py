import os

from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Color
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

file_path = r"C:\Users\nejat.gunaydin\Desktop\Gls-UrgentCargo (002).xlsx"
cmp_colname_1, cmp_colname_2 = 'Title-GSL', 'Title-UrgentCargo'

file_name, ext = os.path.splitext(file_path)
workbook = load_workbook(filename=file_path)
sheetNames = workbook.sheetnames
sheet = workbook[sheetNames[0]]
row_count, column_count = sheet.max_row, sheet.max_column
title_row = [title for title in [value for value in sheet.iter_rows(max_row=1, values_only=True)][0]]
column1_no = title_row.index(cmp_colname_1)+1
column2_no = title_row.index(cmp_colname_2)+1
sheet.cell(row=1, column=column_count+1, value="Title Aynı")
for row in range(2, row_count+1):
    if sheet.cell(row=row, column=column1_no).value != sheet.cell(row=row, column=column2_no).value:
        sheet.cell(row=row, column=column_count + 1, value="Hayır")
        for col in range(1, column_count + 2):
            sheet.cell(row=row, column=col).fill = PatternFill("solid", fgColor=colors.RED)
    else:
        sheet.cell(row=row, column=column_count + 1, value="Evet")

workbook.save(filename=file_name+"-compared-colored"+ext)
